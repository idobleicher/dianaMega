"""Build one consolidated, formatted UBR3 deliverable workbook."""
import shutil
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

cls = 'UBR3_substrates_classified.xlsx'
enr = 'UBR3_2ndAA_enrichment.xlsx'
A   = pd.read_excel(cls, sheet_name='A_rescue_confirmed')
B   = pd.read_excel(cls, sheet_name='B_broad_KOvsAAVS')
S   = pd.read_excel(cls, sheet_name='Summary_by_2ndAA')
E2  = pd.read_excel(enr, sheet_name='enrich_2nd_residue')
EE  = pd.read_excel(enr, sheet_name='enrich_effective_Nterm')

legend = pd.DataFrame({
 'Item': [
  'EXPERIMENT', 'Design', 'UBR3 levels',
  '', 'SUBSTRATE LOGIC',
  '', 'SHEET: 1_README', 'SHEET: 2_rescue_confirmed', 'SHEET: 3_broad_KOvsAAVS',
  'SHEET: 4_summary_by_2ndAA', 'SHEET: 5_enrich_2nd_residue', 'SHEET: 6_enrich_effective_Nterm',
  '', 'KEY COLUMNS', 'SecondAA_pos2', 'Nterm_pos1', 'Met_excised_pred', 'Effective_Nterm',
  'mean_*_log2', 'Difference UBR3 KO_AAVS', 'Difference UBR3 KO_UBR3 WT',
  'p-value / q-value', 'N.Sequences', 'rescue_confirmed',
  '', 'STATS', 'Filter', 'Enrichment test', 'HEADLINE RESULT',
  '', 'NOTES', 'Imputation', 'Clonal confound'],
 'Explanation': [
  'UBR3 substrate screen by quantitative MS (DIA), 3 lines x 3 replicates',
  'AAVS = safe-harbor control (UBR3 present) | UBR3 KO (UBR3 absent) | KO + UBR3 WT = rescue (UBR3 restored)',
  'AAVS 22.5 (present), KO 14.0 (=floor, knocked out), WT 24.6 (re-expressed)',
  '', 'UBR3 is present in BOTH AAVS and WT-rescue, absent only in KO. A real substrate accumulates in KO vs BOTH and drops in the rescue.',
  '', 'This legend',
  'PRIMARY list (n=501): KO>AAVS AND KO>WT, both t-test p<0.05, >=2 peptides. Trustworthy substrates.',
  'BROAD list (n=2381): your literal criterion KO>AAVS, p<0.05, >=2 pep. Only 501 (21%) are rescue-confirmed; rest = likely clonal.',
  'Count of each 2nd amino acid across List A vs List B',
  'Fisher exact enrichment of each 2nd residue (501 substrates vs 8222 detected proteome), BH-FDR corrected',
  'Same test on the effective N-terminus after initiator-Met excision',
  '', '', '2nd amino acid = residue at position 2 of the canonical UniProt sequence (what you asked to classify by)',
  'N-terminal residue (position 1); almost always Met',
  'yes if pos1=Met and pos2 in {A,C,G,P,S,T,V} -> initiator Met is removed, pos2 becomes the new N-terminus',
  'Residue actually exposed at the N-terminus (=pos2 if Met excised, else Met)',
  'mean log2 intensity per group (AAVS / WT / KO); 14.0 = not-detected imputation floor',
  'log2 fold change KO minus AAVS (positive = higher in KO)',
  'log2 fold change KO minus WT-rescue (positive = higher in KO, i.e. rescued back down)',
  "Student's t-test p (raw) and q (FDR across the 8731 proteins)",
  'number of peptides/sequences supporting the protein (>=2 required)',
  'yes = the rescue agrees (KO>WT, p<0.05); no = KO>AAVS only',
  '', '', 'p<0.05 + >=2 peptides + positive accumulation (no fold-change floor, per request)',
  'Two-sided Fisher exact per residue, Benjamini-Hochberg FDR over 20 residues',
  'NO 2nd residue is significantly enriched/depleted after FDR (best q~0.26). Ala dominance = base rate, not enrichment. Nominal only: Leu/Asp/Lys up, Pro/Thr/His down.',
  '', '', 'Value 14.0 = constant imputation floor for not-detected proteins',
  'KO and rescue share clonal background; AAVS is a separate clone, so KO-vs-AAVS alone is confounded -> use rescue-confirmed list.'],
})

out = 'UBR3_substrate_analysis.xlsx'
A2 = A.rename(columns={'SecondAA_pos2':'2nd_AA','Nterm_pos1':'Nterm_pos1','Effective_Nterm':'Effective_Nterm'})
with pd.ExcelWriter(out, engine='openpyxl') as xw:
    legend.to_excel(xw, '1_README', index=False)
    A.to_excel(xw, '2_rescue_confirmed', index=False)
    B.to_excel(xw, '3_broad_KOvsAAVS', index=False)
    S.to_excel(xw, '4_summary_by_2ndAA', index=False)
    E2.to_excel(xw, '5_enrich_2nd_residue', index=False)
    EE.to_excel(xw, '6_enrich_effective_Nterm', index=False)

    wb = xw.book
    hdr_fill = PatternFill('solid', fgColor='1F4E78')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        # column widths
        for c in range(1, ws.max_column + 1):
            col = get_column_letter(c)
            maxlen = max([len(str(ws.cell(r, c).value)) if ws.cell(r, c).value is not None else 0
                          for r in range(1, min(ws.max_row, 200) + 1)] + [8])
            ws.column_dimensions[col].width = min(max(maxlen + 2, 10), 70)
        ws.row_dimensions[1].height = 28
    # README wider explanation column
    rm = wb['1_README']; rm.column_dimensions['A'].width = 26; rm.column_dimensions['B'].width = 115

print('wrote', out)
# copy next to the source file in Downloads
dst = r'C:\Users\User\Downloads\UBR3_substrate_analysis.xlsx'
shutil.copy(out, dst)
print('copied to', dst)
