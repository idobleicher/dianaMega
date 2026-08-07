#!/usr/bin/env python
"""Build the reviewer-facing Excel workbook for the UBR3 N24mer screen."""
import os

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import statsmodels.api as sm
from scipy import stats
from scipy.signal import argrelextrema
from scipy.stats import gaussian_kde
from statsmodels.stats.multitest import multipletests

import ubr3_core as U

OUT = os.path.join(U.HERE, 'UBR3_PG_substrate_tables.xlsx')

ID_COLS = ['rank', 'Gene_ID', 'ENST_ID', 'peptide_24mer']
NT_COLS = ['pos1', 'pos2', 'pos3', 'Met_excised_pred', 'Nterm_after_MetAP',
           'Nterm_motif', 'motif_class']
ANN_COLS = ['uniprot', 'protein_name', 'function', 'localization',
            'go_biological_process', 'protein_length']
QUANT = ['mean_PSI_control', 'mean_PSI_UBR3KO', 'PSI_class_control', 'PSI_class_UBR3KO',
         'crosses_PSI3_up', U.D1, U.D2, 'mean_dPSI',
         'min_dPSI', 'rep_agreement', 'min_total_reads']
PROP = ['net_charge_24mer', 'hydropathy_GRAVY', 'acidic_count_24mer', 'basic_count_24mer']


def fisher_table(hit_counts, bg_counts, labels, label_name):
    """Two-sided Fisher exact per label, hits vs the rest of the library, BH-FDR."""
    nh, nb = hit_counts.sum(), bg_counts.sum()
    rows = []
    for lab in labels:
        a, b = hit_counts.get(lab, 0), bg_counts.get(lab, 0)
        table = [[a, nh - a], [b, nb - b]]
        odds, p = stats.fisher_exact(table)
        fh, fb = a / nh if nh else 0, b / nb if nb else 0
        rows.append({label_name: lab, 'n_hits': int(a), 'pct_hits': round(100 * fh, 2),
                     'n_library': int(b), 'pct_library': round(100 * fb, 2),
                     'fold_enrichment': round(fh / fb, 2) if fb else np.nan,
                     'log2_fold': round(np.log2(fh / fb), 3) if fb and fh else np.nan,
                     'odds_ratio': round(odds, 3) if np.isfinite(odds) else np.nan,
                     'p_value': p})
    df = pd.DataFrame(rows)
    df['q_value_BH'] = multipletests(df.p_value, method='fdr_bh')[1]
    df['significant_q<0.05'] = np.where(df['q_value_BH'] < 0.05, 'yes', 'no')
    df['p_value'] = df.p_value.map(lambda x: float(f'{x:.3g}'))
    df['q_value_BH'] = df.q_value_BH.map(lambda x: float(f'{x:.3g}'))
    return df.sort_values('p_value')


def readme_sheet(stats_d):
    s = stats_d
    CUT = U.PSI_CUT
    rows = [
        ('SOURCE', 'Supplemental Data 1.xlsx - sheet (A) N24mer library, sheet (B) UBR3 substrates'),
        ('ASSAY', 'GPS (Global Protein Stability): each N-terminal 24-mer is fused to a fluorescent '
                  'reporter, cells are FACS-sorted into 4 stability bins, and bin distribution is read '
                  'out by Illumina sequencing.'),
        ('PSI', 'Protein Stability Index - read-count-weighted mean bin (range ~1-4). Higher = more stable.'),
        ('dPSI', 'PSI(UBR3 KO) - PSI(control KO). Positive = the peptide is stabilised when UBR3 is lost, '
                 'i.e. UBR3 normally destabilises it.'),
        ('REPLICATES', 'Two independent control-KO and two independent UBR3-KO replicates.'),
        ('', ''),
        ('SUBSTRATE DEFINITION', ''),
        ('The substrate set', f'The {s["n_hits"]} peptides in sheet (B) - the final candidate UBR3 '
                              'substrates - are the "stabilised" set everywhere in this workbook and in '
                              'every figure. Every peptide in the library that is not one of these '
                              f'{s["n_hits"]} is treated as not stabilised.'),
        ('', ''),
        ('N-TERMINAL LOGIC', ''),
        ('pos1 / pos2 / pos3', 'Residues 1-3 of the encoded 24-mer. pos1 is Met in 100% of the library.'),
        ('Met_excised_pred', 'yes when pos1 = Met and pos2 is in {A,C,G,P,S,T,V} (Sherman\'s rule) - MetAP '
                             'removes the initiator Met, so pos2 becomes the true N-terminal residue.'),
        ('Nterm_after_MetAP', 'The residue actually exposed at the N-terminus (= pos2 if excised, else Met).'),
        ('[P/G]-[E/D]', 'pos2 in {P,G} AND pos3 in {E,D}. After Met excision this is a Pro- or Gly-initiated '
                        'N-terminus immediately followed by an acidic residue - the motif under test.'),
        ('[P/G]-other', 'pos2 in {P,G} but pos3 NOT acidic.'),
        ('non-P/G', 'pos2 is neither Pro nor Gly.'),
        ('CAVEAT', 'Met excision at Xaa-Pro is less efficient than at Xaa-Gly; the predicted N-terminus for '
                   'Pro-starting peptides should be read as a prediction, not a measurement.'),
        ('', ''),
        ('DATA REPAIR: GENE SYMBOLS', '36 gene symbols in sheet (A) of the source file are stored as DATES '
                                      '(e.g. 2005-03-01), the well-known Excel corruption of MARCH and SEPT '
                                      'family symbols - MARCH5 becomes 1-Mar, SEPT2 becomes 1-Sep. All 36 '
                                      'have been restored here from their Ensembl transcript IDs, which are '
                                      'not corrupted. Column "gene_symbol_repaired" = yes flags every '
                                      'repaired row. NONE of the 54 substrates are affected, so no result in '
                                      'this workbook changes - but the source supplementary file should be '
                                      'corrected before publication.'),
        ('', ''),
        ('HOW SUBSTRATES ARE MARKED', 'On every peptide listing that mixes substrates with non-substrates '
                                      '(sheets 02 and 03) the ENTIRE ROW of a UBR3 substrate is shaded '
                                      'orange and bolded, so the handful of substrates are visible while '
                                      'scrolling thousands of rows. The column "UBR3_substrate" also reads '
                                      'yes/no, and every sheet is auto-filtered - filter it to yes to '
                                      'isolate them. Sheet 01 is substrates only, so nothing is highlighted '
                                      'there.'),
        ('', ''),
        ('HEADLINE NUMBERS (sheet 05)', ''),
        ('Library size', f'{s["n_lib"]:,} N24mer peptides, all 24 aa long, all starting with Met.'),
        ('P/G at pos2', f'{s["n_pg"]:,} peptides ({s["pct_pg"]:.2f}% of library) vs {s["n_hit_pg"]} of '
                        f'{s["n_hits"]} substrates ({s["pct_hit_pg"]:.1f}%).'),
        ('[P/G]-[E/D] peptides', f'{s["n_pged"]} in the library; {s["n_hit_pged"]} are UBR3-stabilised '
                                 f'({s["rate_pged"]:.2f}%).'),
        ('CENTRAL CLAIM', f'{s["n_pged"] - s["n_hit_pged"]} of {s["n_pged"]} [P/G]-[E/D] peptides '
                          f'({100 - s["rate_pged"]:.1f}%) show NO stabilisation on UBR3 loss. The motif is '
                          'therefore necessary-but-not-sufficient: it enriches for UBR3 substrates '
                          f'{s["fold_pged_vs_non"]:.0f}-fold over non-P/G peptides '
                          f'({s["rate_pged"]:.2f}% vs {s["rate_non"]:.2f}%) but only a small subset of '
                          'motif-bearing peptides is actually UBR3-regulated.'),
        ('', ''),
        ('SHEETS', ''),
        ('01_UBR3_substrates', f'All {s["n_hits"]} candidate substrates, ranked by mean dPSI, with UniProt '
                               'functional annotation. This is the clean readable table the reviewer asked for.'),
        ('02_PGED_motif_all', f'THE KEY REVIEWER TABLE. All {s["n_pged"]} library peptides carrying '
                              '[P/G]-[E/D], stabilised and non-stabilised side by side, sorted by mean dPSI.'),
        ('03_PG_library_all', f'All {s["n_pg"]:,} peptides with Pro or Gly at pos2, sorted by mean dPSI.'),
        ('04_substrates_PG_only', f'The {s["n_hit_pg"]} substrates that carry P/G at pos2 (subset of sheet 01).'),
        ('05_counts_summary', 'Every count behind the motif-class comparisons, as a flat table.'),
        ('06_enrich_pos2_residue', 'Fisher exact enrichment of each of the 20 residues at pos2, substrates '
                                   'vs library, BH-FDR corrected.'),
        ('07_enrich_motif_class', 'Same test at the level of the three motif classes and of pos3 identity.'),
        ('08_AAclass_by_position', 'Fraction of each chemical class at each of the 24 positions, substrates '
                                   'vs library, plus log2 enrichment. Drives the class heatmap and bar charts.'),
        ('09_posAA_freq_substrates', '24 x 20 position-frequency matrix for the substrates (logo input).'),
        ('10_posAA_freq_library', 'Same matrix for the full library (logo background).'),
        ('11_posAA_log2_enrichment', 'log2[ substrate freq / library freq ] per position per residue - '
                                     'the enrichment heatmap and the enrichment logo.'),
        ('12_posAA_fisher_all_cells', 'All 480 position x residue cells tested by Fisher exact with '
                                      'BH-FDR across the whole matrix. Sorted by q. Only 4 cells reach '
                                      'q<0.05: Asp at position 3, Pro and Gly at position 2, and Arg at '
                                      'position 7. This is the table behind the FDR-masked heatmaps.'),
        ('13_posAA_qvalues', 'The same q values as a 24 x 20 matrix for quick lookup.'),
        ('14_PSI_strata_summary', 'Splits the library at control PSI = 3 into "unstable at baseline" and '
                                  '"already stable" and reports size, substrate rate, mean dPSI and motif '
                                  'content for each, plus how many peptides cross the line on UBR3 loss.'),
        ('15_motif_x_PSI_strata', 'THE KEY CONTROL. Substrate rate for each motif class WITHIN each '
                                  'baseline-stability stratum, with a Fisher test per stratum. Shows the '
                                  'motif effect is not a by-product of the peptides already being unstable.'),
        ('16_substrates_PSI_transition', 'Each of the 54 substrates with its control PSI, UBR3-KO PSI, '
                                         'stratum on each side and whether it crosses PSI = 3 upward.'),
        ('17_dPSI_by_control_PSI_bin', 'dPSI, substrate count and crossing rate in 0.5-wide bins of '
                                       'control PSI - the ceiling effect, in numbers.'),
        ('19_stability_x_substrate', f'THE CLASSIFICATION TABLE. Every motif group crossed with the '
                                     f'stability call at control PSI {CUT:g} and with substrate status: how '
                                     'many peptides are unstable vs stable, how many of each are UBR3 '
                                     'substrates, and the Fisher test comparing the two strata. NOTE: the '
                                     'unstable-vs-stable contrast is significant library-wide (OR 2.97, '
                                     'p = 3e-4) but NOT within [P/G]-[E/D] alone (OR 2.85, p = 0.067, only '
                                     '11 vs 5 substrates). Within the motif class treat it as directional.'),
        ('20_PGED_substrates_annotated', f'The {s["n_hit_pged"]} [P/G]-[E/D] peptides that ARE UBR3 '
                                         'substrates, split by stability class and fully annotated '
                                         '(UniProt accession, protein name, function, localisation, GO), '
                                         'with control PSI, UBR3-KO PSI and whether each crosses the line.'),
        ('21_downstream_residue', 'THE DOWNSTREAM COMPARISON. Positions 4-24 only, comparing the 16 motif-bearing SUBSTRATES with the 163 motif-bearing NON-substrates. Both groups carry M-[P/G]-[E/D], so positions 1-3 are identical by construction and the motif itself cannot drive any difference. Fisher exact per position x residue with BH-FDR over all 420 cells. RESULT: nothing survives; only 11 cells reach nominal p<0.05, fewer than the ~21 expected by chance.'),
        ('22_downstream_class', 'Same comparison at the level of the six chemical classes (126 cells). One survives FDR: Basic at position 5 (50% of substrates vs 9.8% of non-substrates, q = 0.026).'),
        ('23_downstream_composition', 'The better-powered aggregate view: per-peptide counts of each class summed over positions 4-24, plus net charge and GRAVY, compared by Mann-Whitney. RESULT: substrates carry more basic residues (4.19 vs 2.91, p = 0.014) and a markedly more positive net charge (+2.16 vs -0.22, p = 0.0021). Nothing else differs.'),
        ('24_PG_substrate_class_composition', 'Chemical-class composition at each of the 24 positions for three groups side by side: the 26 substrates with Pro/Gly at position 2, the 2,100 P/G library peptides (the motif-matched background) and the whole library. Also gives log2 enrichment and FDR q of the P/G substrates against the P/G background - only Acidic at position 3 survives.'),
        ('25_PGsub_vs_PGnonsub_class', 'The cleanest motif-controlled comparison: the 26 P/G peptides that ARE substrates against the 2,074 P/G peptides that are NOT. Class composition per position for both, with log2 enrichment and FDR q. Only Acidic at position 3 survives.'),
        ('26_PGsub_vs_PGnonsub_residue', 'The same comparison per individual residue, all 480 position x residue cells sorted by q. Three survive: Asp at position 3 (q = 5e-5), Glu at position 3 (q = 0.034) and Arg at position 7 (q = 0.034).'),
        ('18_PSI_cutoff_robustness', 'Shows the PSI = 3 cut is not doing the work: the motif odds ratio '
                                     'across cutoffs 2.0-3.4, a logistic regression using control PSI as a '
                                     'CONTINUOUS covariate (no cutoff at all), the peaks and valleys of the '
                                     'control-PSI density, and the Spearman correlation between control PSI '
                                     'and dPSI.'),
        ('', ''),
        ('WHY NOT SPLIT ON dPSI', 'dPSI measures UBR3 DEPENDENCE, not stability - and it is the variable the '
                                  '54 substrates were selected on, so using it to define the strata and then '
                                  'asking where the substrates fall is circular (every dPSI cut contains all '
                                  '54). Control PSI measures the stability the peptide has BEFORE UBR3 is '
                                  'removed. The two are nearly independent (Spearman rho = -0.20, ~3% shared '
                                  'variance), which is why stratifying on control PSI is a genuine control '
                                  'and stratifying on dPSI is not.'),
        ('', ''),
        ('BASELINE STABILITY (PSI) ANALYSIS', ''),
        (f'Why split at PSI = {CUT:g}', 'PSI runs ~1-4 (read-weighted mean FACS bin). The control-PSI '
                                        'density has three robust modes (~1.52, ~2.25, ~3.49); the antimode '
                                        'between the unstable modes and the stable one sits at 2.59-2.63 '
                                        f'across kernel bandwidths, so the cut is placed at {CUT:g}. It '
                                        f'splits the library into {s["n_lo"]:,} unstable ({s["pct_lo"]:.1f}%) '
                                        f'and {s["n_hi"]:,} already-stable peptides. Sheet 18 shows nothing '
                                        'depends on this choice.'),
        ('Substrates start unstable', f'{s["n_hit_lo"]} of the {s["n_hits"]} substrates ({s["pct_hit_lo"]:.1f}%) '
                                      f'have control PSI < {CUT:g}. Substrate rate is {s["rate_lo"]:.3f}% below '
                                      f'the line vs {s["rate_hi"]:.3f}% above it '
                                      f'({s["rate_lo"] / s["rate_hi"]:.1f}-fold). NOTE: this particular '
                                      'contrast is cutoff-sensitive (it is ~16-fold at PSI 3) because the '
                                      'assay ceiling suppresses detectable stabilisation at high PSI; the '
                                      'motif result in sheet 15 is not.'),
        ('Substrates cross the line', f'{s["n_cross"]} of the {s["n_hit_lo"]} substrates starting below PSI '
                                      f'{CUT:g} ({s["pct_cross"]:.1f}%) rise above it when UBR3 is lost, '
                                      f'against a library background of {s["pct_cross_lib"]:.1f}%.'),
        ('Ceiling effect', 'Mean dPSI peaks in the 2.5-3.0 bin (+0.150) and turns negative above 3.5 '
                           '(-0.067): peptides that are already maximally stable cannot be stabilised '
                           'further. Sheet 17.'),
        ('The motif is not a stability artefact', f'[P/G]-[E/D] is carried by {s["pged_lo"]:.2f}% of unstable '
                                                  f'and {s["pged_hi"]:.2f}% of stable peptides - i.e. it is '
                                                  'NOT over-represented among intrinsically unstable '
                                                  f'peptides. Fisher OR for [P/G]-[E/D] vs non-P/G is '
                                                  f'{s["or_lo"]:.0f} within the unstable stratum and '
                                                  f'{s["or_hi"]:.0f} within the stable one - essentially '
                                                  'identical, so there is no interaction with baseline '
                                                  'stability. Sheet 15.'),
        ('', ''),
        ('AMINO ACID CLASSES', 'Acidic D,E | Basic K,R,H | Polar S,T,N,Q,C | Hydrophobic A,V,L,I,M | '
                               'Aromatic F,W,Y | Special G,P'),
        ('STATISTICS', 'Two-sided Fisher exact test; Benjamini-Hochberg FDR across the residues/classes '
                       'tested in each sheet. Group dPSI comparisons use the two-sided Mann-Whitney U test.'),
    ]
    return pd.DataFrame(rows, columns=['Item', 'Explanation'])


def main():
    lib, hit = U.load()
    lib_seqs, hit_seqs = list(lib.peptide_24mer), list(hit.peptide_24mer)

    # ---------- headline counts ----------
    n_lib, n_hits = len(lib), len(hit)
    pged = lib[lib.motif_class == '[P/G]-[E/D]']
    pgo = lib[lib.motif_class == '[P/G]-other']
    non = lib[lib.motif_class == 'non-P/G']
    rate_pged = 100 * pged.is_UBR3_substrate.mean()
    rate_non = 100 * non.is_UBR3_substrate.mean()
    S = {'n_lib': n_lib, 'n_hits': n_hits,
         'n_pg': int(lib.is_PG.sum()), 'pct_pg': 100 * lib.is_PG.mean(),
         'n_hit_pg': int(hit.is_PG.sum()), 'pct_hit_pg': 100 * hit.is_PG.mean(),
         'n_pged': len(pged), 'n_hit_pged': int(pged.is_UBR3_substrate.sum()),
         'rate_pged': rate_pged, 'rate_non': rate_non,
         'fold_pged_vs_non': rate_pged / rate_non}
    _lo = lib.mean_PSI_control < U.PSI_CUT
    _hlo = hit.mean_PSI_control < U.PSI_CUT
    S.update({
        'n_lo': int(_lo.sum()), 'n_hi': int((~_lo).sum()), 'pct_lo': 100 * _lo.mean(),
        'n_hit_lo': int(_hlo.sum()), 'pct_hit_lo': 100 * _hlo.mean(),
        'rate_lo': 100 * lib[_lo].is_UBR3_substrate.mean(),
        'rate_hi': 100 * lib[~_lo].is_UBR3_substrate.mean(),
        'n_cross': int((hit.crosses_PSI3_up == 'yes').sum()),
        'pct_cross': 100 * (hit.crosses_PSI3_up == 'yes').sum() / max(int(_hlo.sum()), 1),
        'pct_cross_lib': 100 * (lib[_lo].crosses_PSI3_up == 'yes').mean(),
        'pged_lo': 100 * lib[_lo].is_PG_ED.mean(), 'pged_hi': 100 * lib[~_lo].is_PG_ED.mean()})
    for key, mask in [('or_lo', _lo), ('or_hi', ~_lo)]:
        _s = lib[mask]
        _a, _b = _s[_s.motif_class == '[P/G]-[E/D]'], _s[_s.motif_class == 'non-P/G']
        S[key] = stats.fisher_exact([[int(_a.is_UBR3_substrate.sum()), int((~_a.is_UBR3_substrate).sum())],
                                     [int(_b.is_UBR3_substrate.sum()), int((~_b.is_UBR3_substrate).sum())]])[0]

    # ---------- sheet 01 : annotated substrates ----------
    s01 = hit[ID_COLS + NT_COLS + ANN_COLS + QUANT + PROP + ['peptide_24mer']].copy()
    s01 = s01.loc[:, ~s01.columns.duplicated()]

    # ---------- sheet 02 : the reviewer's [P/G]-[E/D] table ----------
    key = ['Gene_ID', 'gene_symbol_repaired', 'ENST_ID', 'peptide_24mer', 'pos2', 'pos3',
           'Nterm_motif', 'Nterm_after_MetAP', 'Met_excised_pred', 'UBR3_substrate',
           'stability_class', 'classification'] + QUANT + PROP
    s02 = pged[key].copy().sort_values('mean_dPSI', ascending=False)
    ann_map = hit.set_index('Gene_ID')[['protein_name', 'function']].to_dict('index')
    s02['protein_name'] = [ann_map.get(g, {}).get('protein_name', '') for g in s02.Gene_ID]
    s02['function'] = [ann_map.get(g, {}).get('function', '') for g in s02.Gene_ID]
    s02.insert(0, 'rank_by_dPSI', np.arange(1, len(s02) + 1))

    # ---------- sheet 03 / 04 ----------
    s03 = lib[lib.is_PG][key].copy().sort_values('mean_dPSI', ascending=False)
    s03.insert(0, 'rank_by_dPSI', np.arange(1, len(s03) + 1))
    s04 = s01[s01.motif_class != 'non-P/G'].copy()

    # ---------- sheet 05 : counts ----------
    rows = [('Library', 'total N24mer peptides', n_lib, ''),
            ('Library', 'peptides with Met at pos1', int((lib.pos1 == 'M').sum()), '100%'),
            ('Substrates', 'candidate UBR3 substrates (sheet B)', n_hits, '')]
    for label, sub in [('P/G at pos2', lib[lib.is_PG]), ('Pro at pos2', lib[lib.pos2 == 'P']),
                       ('Gly at pos2', lib[lib.pos2 == 'G']),
                       ('[P/G]-[E/D]', pged), ('[P/G]-other', pgo), ('non-P/G', non)]:
        rows.append(('Motif group', f'{label}: library peptides', len(sub),
                     f'{100 * len(sub) / n_lib:.2f}% of library'))
        rows.append(('Motif group', f'{label}: UBR3-stabilised', int(sub.is_UBR3_substrate.sum()),
                     f'{100 * sub.is_UBR3_substrate.mean():.2f}% of this group'))
        rows.append(('Motif group', f'{label}: NOT stabilised', int((~sub.is_UBR3_substrate).sum()),
                     f'{100 * (~sub.is_UBR3_substrate).mean():.2f}% of this group'))
    for lab, sub in [('P', pged[pged.pos2 == 'P']), ('G', pged[pged.pos2 == 'G'])]:
        rows.append(('[P/G]-[E/D] split', f'{lab}-[E/D] library / stabilised', len(sub),
                     f'{int(sub.is_UBR3_substrate.sum())} stabilised '
                     f'({100 * sub.is_UBR3_substrate.mean():.1f}%)'))
    # Mann-Whitney between groups
    for a, b in [('[P/G]-[E/D]', '[P/G]-other'), ('[P/G]-[E/D]', 'non-P/G'), ('[P/G]-other', 'non-P/G')]:
        x = lib[lib.motif_class == a].mean_dPSI
        y = lib[lib.motif_class == b].mean_dPSI
        u, p = stats.mannwhitneyu(x, y, alternative='two-sided')
        rows.append(('Statistics', f'mean dPSI {a} vs {b} (Mann-Whitney U)', float(f'{p:.3g}'),
                     f'medians {x.median():.4f} vs {y.median():.4f}'))
    ct = [[int(pged.is_UBR3_substrate.sum()), int((~pged.is_UBR3_substrate).sum())],
          [int(non.is_UBR3_substrate.sum()), int((~non.is_UBR3_substrate).sum())]]
    orv, pv = stats.fisher_exact(ct)
    rows.append(('Statistics', '[P/G]-[E/D] vs non-P/G stabilisation rate (Fisher exact)',
                 float(f'{pv:.3g}'), f'odds ratio {orv:.1f}'))
    s05 = pd.DataFrame(rows, columns=['Category', 'Metric', 'Value', 'Note'])

    # ---------- sheet 06 / 07 : enrichment ----------
    s06 = fisher_table(hit.pos2.value_counts(), lib.pos2.value_counts(), U.AAS, 'residue_at_pos2')
    s07a = fisher_table(hit.motif_class.value_counts(), lib.motif_class.value_counts(),
                        ['[P/G]-[E/D]', '[P/G]-other', 'non-P/G'], 'group')
    s07b = fisher_table(hit.pos3.value_counts(), lib.pos3.value_counts(), U.AAS, 'group')
    s07b['group'] = 'pos3=' + s07b['group']
    s07c = fisher_table(hit.Nterm_after_MetAP.value_counts(), lib.Nterm_after_MetAP.value_counts(),
                        sorted(lib.Nterm_after_MetAP.unique()), 'group')
    s07c['group'] = 'Nterm_after_MetAP=' + s07c['group']
    s07 = pd.concat([s07a, s07b, s07c], ignore_index=True)

    # ---------- sheet 08 : class by position ----------
    ch, cl = U.class_matrix(hit_seqs), U.class_matrix(lib_seqs)
    s08 = pd.DataFrame({'position': ch.index})
    for c in U.CLASS_ORDER:
        s08[f'{c}_substrates_pct'] = (100 * ch[c]).round(2).values
        s08[f'{c}_library_pct'] = (100 * cl[c]).round(2).values
        s08[f'{c}_log2_enrich'] = np.log2((ch[c].values + 1e-9) / (cl[c].values + 1e-9)).round(3)

    # ---------- sheet 09-11 : position x AA ----------
    fh, fl = U.freq_matrix(hit_seqs), U.freq_matrix(lib_seqs)
    s09 = (100 * fh).round(2).rename_axis('position').reset_index()
    s10 = (100 * fl).round(2).rename_axis('position').reset_index()
    pseudo = 1.0 / len(hit_seqs)
    s11 = np.log2((fh + pseudo) / (fl + pseudo)).round(3).rename_axis('position').reset_index()

    # ---------- sheet 12/13 : per-cell Fisher across all 24 x 20 cells -------
    lr, qv, _ = U.enrichment_test(hit_seqs, lib_seqs)
    ch_cnt, cl_cnt = U.position_matrix(hit_seqs), U.position_matrix(lib_seqs)
    rows = []
    for pos in qv.index:
        for a in qv.columns:
            rows.append({'position': pos, 'residue': a, 'chemical_class': U.AA_CLASS[a],
                         'n_substrates': int(ch_cnt.loc[pos, a]),
                         'pct_substrates': round(100 * ch_cnt.loc[pos, a] / len(hit_seqs), 2),
                         'n_library': int(cl_cnt.loc[pos, a]),
                         'pct_library': round(100 * cl_cnt.loc[pos, a] / len(lib_seqs), 2),
                         'log2_enrichment': round(lr.loc[pos, a], 3),
                         'q_value_BH': float(f'{qv.loc[pos, a]:.3g}'),
                         'significant_q<0.05': 'yes' if qv.loc[pos, a] < 0.05 else 'no'})
    s12 = pd.DataFrame(rows).sort_values('q_value_BH')
    s13 = qv.round(4).rename_axis('position').reset_index()

    # ---------- sheet 14-17 : baseline-stability (PSI) strata ---------------
    CUT = U.PSI_CUT
    lo = lib.mean_PSI_control < CUT
    rows = []
    for lab, m in [(f'control PSI < {CUT} (unstable at baseline)', lo),
                   (f'control PSI >= {CUT} (already stable)', ~lo)]:
        s = lib[m]
        rows += [('Library split', f'{lab}: peptides', len(s), f'{100 * len(s) / n_lib:.1f}% of library'),
                 ('Library split', f'{lab}: UBR3 substrates', int(s.is_UBR3_substrate.sum()),
                  f'{100 * s.is_UBR3_substrate.mean():.3f}% of this stratum'),
                 ('Library split', f'{lab}: mean dPSI', round(s.mean_dPSI.mean(), 4),
                  f'median {s.mean_dPSI.median():.4f}'),
                 ('Library split', f'{lab}: carries [P/G]-[E/D]', int(s.is_PG_ED.sum()),
                  f'{100 * s.is_PG_ED.mean():.2f}% of this stratum')]
    for lab, m in [('control', hit.mean_PSI_control < CUT), ('UBR3 KO', hit.mean_PSI_UBR3KO < CUT)]:
        rows.append(('Substrates', f'{lab} PSI < {CUT}', int(m.sum()),
                     f'{100 * m.mean():.1f}% of the {n_hits} substrates'))
    up = (hit.crosses_PSI3_up == 'yes').sum()
    lib_up = (lib[lo].crosses_PSI3_up == 'yes').sum()
    rows += [('Substrates', f'cross PSI {CUT} upward on UBR3 loss', int(up),
              f'{100 * up / (hit.mean_PSI_control < CUT).sum():.1f}% of substrates starting below {CUT}'),
             ('Library', f'cross PSI {CUT} upward on UBR3 loss', int(lib_up),
              f'{100 * lib_up / lo.sum():.1f}% of library peptides starting below {CUT}')]
    orv, pv = stats.fisher_exact([[int(lib[lo].is_UBR3_substrate.sum()),
                                   int((~lib[lo].is_UBR3_substrate).sum())],
                                  [int(lib[~lo].is_UBR3_substrate.sum()),
                                   int((~lib[~lo].is_UBR3_substrate).sum())]])
    rows.append(('Statistics', f'substrate rate PSI<{CUT} vs PSI>={CUT} (Fisher exact)',
                 float(f'{pv:.3g}'), f'odds ratio {orv:.1f}'))
    s14 = pd.DataFrame(rows, columns=['Category', 'Metric', 'Value', 'Note'])

    # the key control: does the motif effect survive stratification?
    rows = []
    for stratum, m in [(f'PSI < {CUT} (unstable)', lo), (f'PSI >= {CUT} (stable)', ~lo)]:
        sub = lib[m]
        for mc in ['[P/G]-[E/D]', '[P/G]-other', 'non-P/G']:
            g = sub[sub.motif_class == mc]
            rows.append({'baseline_stratum': stratum, 'motif_class': mc, 'n_peptides': len(g),
                         'n_substrates': int(g.is_UBR3_substrate.sum()),
                         'substrate_rate_pct': round(100 * g.is_UBR3_substrate.mean(), 3),
                         'median_dPSI': round(g.mean_dPSI.median(), 4),
                         'median_PSI_control': round(g.mean_PSI_control.median(), 3)})
        a, b = sub[sub.motif_class == '[P/G]-[E/D]'], sub[sub.motif_class == 'non-P/G']
        orv, pv = stats.fisher_exact([[int(a.is_UBR3_substrate.sum()), int((~a.is_UBR3_substrate).sum())],
                                      [int(b.is_UBR3_substrate.sum()), int((~b.is_UBR3_substrate).sum())]])
        rows.append({'baseline_stratum': stratum, 'motif_class': '[P/G]-[E/D] vs non-P/G',
                     'n_peptides': None, 'n_substrates': None,
                     'substrate_rate_pct': None, 'median_dPSI': None,
                     'median_PSI_control': f'Fisher OR={orv:.1f}, p={pv:.3g}'})
    s15 = pd.DataFrame(rows)

    s16 = hit[['rank', 'Gene_ID', 'peptide_24mer', 'motif_class', 'Nterm_after_MetAP',
               'mean_PSI_control', 'mean_PSI_UBR3KO', 'PSI_class_control', 'PSI_class_UBR3KO',
               'PSI_transition', 'crosses_PSI3_up', U.D1, U.D2, 'mean_dPSI',
               'protein_name']].copy()

    bins = [0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0]
    lib['_bin'] = pd.cut(lib.mean_PSI_control, bins)
    g = lib.groupby('_bin', observed=True)
    s17 = pd.DataFrame({
        'control_PSI_bin': [str(i) for i in g.size().index],
        'n_peptides': g.size().values,
        'mean_dPSI': g.mean_dPSI.mean().round(4).values,
        'median_dPSI': g.mean_dPSI.median().round(4).values,
        'n_substrates': g.is_UBR3_substrate.sum().values,
        'substrate_rate_pct': (100 * g.is_UBR3_substrate.mean()).round(3).values,
        'pct_crossing_PSI3_up': (100 * g.crosses_PSI3_up.apply(lambda s: (s == 'yes').mean())).round(2).values,
    })
    lib.drop(columns='_bin', inplace=True)

    # ---------- sheet 18 : is the PSI cut arbitrary? ------------------------
    rows = []
    for c in [2.0, 2.2, 2.4, 2.6, 2.8, 3.0, 3.2, 3.4]:
        s = lib[lib.mean_PSI_control < c]
        a, b = s[s.motif_class == '[P/G]-[E/D]'], s[s.motif_class == 'non-P/G']
        orv, pv = stats.fisher_exact([[int(a.is_UBR3_substrate.sum()), int((~a.is_UBR3_substrate).sum())],
                                      [int(b.is_UBR3_substrate.sum()), int((~b.is_UBR3_substrate).sum())]])
        rows.append({'analysis': 'cutoff sensitivity', 'PSI_cutoff': c,
                     'n_unstable_stratum': len(s),
                     'PGED_substrate_rate_pct': round(100 * a.is_UBR3_substrate.mean(), 3),
                     'nonPG_substrate_rate_pct': round(100 * b.is_UBR3_substrate.mean(), 4),
                     'odds_ratio': round(orv, 1), 'p_value': float(f'{pv:.3g}')})
    sens = pd.DataFrame(rows)

    d = lib.copy()
    d['y'] = d.is_UBR3_substrate.astype(int)
    d['PGED'] = (d.motif_class == '[P/G]-[E/D]').astype(int)
    d['PGother'] = (d.motif_class == '[P/G]-other').astype(int)
    Xd = sm.add_constant(d[['PGED', 'PGother', 'mean_PSI_control']])
    fit = sm.Logit(d.y, Xd).fit(disp=0)
    ci = fit.conf_int()
    logit = pd.DataFrame({
        'analysis': 'logistic regression (no cutoff)',
        'term': ['intercept', '[P/G]-[E/D]', '[P/G]-other', 'control PSI (per +1 unit)'],
        'coefficient': fit.params.round(4).values,
        'odds_ratio': np.exp(fit.params).round(3).values,
        'OR_95CI_low': np.exp(ci[0]).round(3).values,
        'OR_95CI_high': np.exp(ci[1]).round(3).values,
        'p_value': [float(f'{p:.3g}') for p in fit.pvalues]})

    # bw 0.25: modal structure is stable here; narrower bandwidths add a shoulder
    # near 2.9 that is gone again by 0.20, so it is not a reproducible mode.
    kde = gaussian_kde(lib.mean_PSI_control, bw_method=0.25)
    gx = np.linspace(1, 4, 900)
    gy = kde(gx)
    vall = gx[argrelextrema(gy, np.less)[0]]
    peaks = gx[argrelextrema(gy, np.greater)[0]]
    shape = pd.DataFrame({'analysis': 'control-PSI density shape',
                          'feature': ['peak'] * len(peaks) + ['valley (antimode)'] * len(vall),
                          'control_PSI': np.round(np.concatenate([peaks, vall]), 3)})
    rho = stats.spearmanr(lib.mean_PSI_control, lib.mean_dPSI)
    orth = pd.DataFrame({'analysis': ['axis independence'],
                         'term': ['control PSI vs mean dPSI'],
                         'coefficient': [round(rho.statistic, 3)],
                         'odds_ratio': [None], 'OR_95CI_low': [None], 'OR_95CI_high': [None],
                         'p_value': [float(f'{rho.pvalue:.3g}')]})
    s18 = pd.concat([sens, logit, shape, orth], ignore_index=True)

    # ---------- sheet 19 : stability x substrate cross-classification --------
    rows = []
    for mc in ['[P/G]-[E/D]', '[P/G]-other', 'non-P/G', 'ALL P/G at pos2', 'WHOLE LIBRARY']:
        grp = {'ALL P/G at pos2': lib[lib.is_PG], 'WHOLE LIBRARY': lib}.get(
            mc, lib[lib.motif_class == mc])
        for stab in ['unstable', 'stable']:
            g = grp[grp.stability_class == stab]
            k = int(g.is_UBR3_substrate.sum())
            rows.append({'group': mc, 'stability_class': f'{stab} (control PSI '
                                                         f'{"<" if stab == "unstable" else ">="} {CUT:g})',
                         'n_peptides': len(g), 'n_UBR3_substrates': k,
                         'n_not_substrates': len(g) - k,
                         'substrate_rate_pct': round(100 * k / len(g), 3) if len(g) else np.nan,
                         'pct_of_group_in_this_stratum': round(100 * len(g) / len(grp), 1),
                         'median_control_PSI': round(g.mean_PSI_control.median(), 3),
                         'median_dPSI': round(g.mean_dPSI.median(), 4)})
        k_all = int(grp.is_UBR3_substrate.sum())
        u, s_ = grp[grp.stability_class == 'unstable'], grp[grp.stability_class == 'stable']
        orv, pv = stats.fisher_exact([[int(u.is_UBR3_substrate.sum()), int((~u.is_UBR3_substrate).sum())],
                                      [int(s_.is_UBR3_substrate.sum()), int((~s_.is_UBR3_substrate).sum())]])
        rows.append({'group': mc, 'stability_class': 'BOTH (total)', 'n_peptides': len(grp),
                     'n_UBR3_substrates': k_all, 'n_not_substrates': len(grp) - k_all,
                     'substrate_rate_pct': round(100 * k_all / len(grp), 3),
                     'pct_of_group_in_this_stratum': 100.0,
                     'median_control_PSI': round(grp.mean_PSI_control.median(), 3),
                     'median_dPSI': round(grp.mean_dPSI.median(), 4),
                     'unstable_vs_stable_Fisher': f'OR={orv:.2f}, p={pv:.3g}'})
    s19 = pd.DataFrame(rows)

    # ---------- sheet 20 : the motif-bearing substrates, annotated ----------
    ann = hit.set_index('peptide_24mer')[['uniprot', 'protein_name', 'function',
                                          'localization', 'go_biological_process']].to_dict('index')
    mm = lib[(lib.motif_class == '[P/G]-[E/D]') & lib.is_UBR3_substrate].copy()
    mm = mm.sort_values(['stability_class', 'mean_dPSI'], ascending=[True, False])
    for c in ['uniprot', 'protein_name', 'function', 'localization', 'go_biological_process']:
        mm[c] = [ann.get(p, {}).get(c, '') for p in mm.peptide_24mer]
    s20 = mm[['Gene_ID', 'uniprot', 'ENST_ID', 'peptide_24mer', 'Nterm_motif',
              'Nterm_after_MetAP', 'stability_class', 'classification',
              'mean_PSI_control', 'mean_PSI_UBR3KO', 'crosses_PSI3_up',
              U.D1, U.D2, 'mean_dPSI', 'protein_name', 'function', 'localization',
              'go_biological_process']].copy()
    s20.insert(0, 'rank_by_dPSI', np.arange(1, len(s20) + 1))

    # ---------- sheet 21-23 : downstream of the motif ------------------------
    mA = list(mm.peptide_24mer)
    mB = list(lib[(lib.motif_class == '[P/G]-[E/D]') & ~lib.is_UBR3_substrate].peptide_24mer)
    s21, s22, pep = U.downstream_compare(mA, mB)
    s21 = s21.rename(columns={'group': 'residue'}).sort_values('p_value')
    s22 = s22.rename(columns={'group': 'chemical_class'}).sort_values('p_value')

    rows = []
    A = pep[pep.group == 'UBR3 substrate']
    B = pep[pep.group == 'not a substrate']
    for col, lab in ([(f'n_{c}', f'{c} residues') for c in U.CLASS_ORDER] +
                     [('net_charge', 'net charge'), ('GRAVY', 'GRAVY hydropathy')]):
        u, pv = stats.mannwhitneyu(A[col], B[col], alternative='two-sided')
        rows.append({'measure': f'{lab}, positions {U.DOWN_FROM}-{U.DOWN_TO}',
                     'substrates_mean': round(A[col].mean(), 3),
                     'substrates_median': round(A[col].median(), 3),
                     'non_substrates_mean': round(B[col].mean(), 3),
                     'non_substrates_median': round(B[col].median(), 3),
                     'difference': round(A[col].mean() - B[col].mean(), 3),
                     'MannWhitney_p': float(f'{pv:.4g}'),
                     'significant_p<0.05': 'yes' if pv < 0.05 else 'no'})
    s23 = pd.DataFrame(rows).sort_values('MannWhitney_p')


    # ---------- sheet 24 : P/G substrate class composition -------------------
    pgS = list(hit[hit.is_PG].peptide_24mer)
    pgL = list(lib[lib.is_PG].peptide_24mer)
    cS, cL, cAll = U.class_matrix(pgS), U.class_matrix(pgL), U.class_matrix(lib_seqs)
    lrp, qvp, _ = U.enrichment_test(pgS, pgL, groups=U.CLASS_MEMBERS)
    s24 = pd.DataFrame({'position': cS.index})
    for c in U.CLASS_ORDER:
        s24[c + '_PGsubstrates_pct'] = (100 * cS[c]).round(2).values
        s24[c + '_PGlibrary_pct'] = (100 * cL[c]).round(2).values
        s24[c + '_wholeLibrary_pct'] = (100 * cAll[c]).round(2).values
        s24[c + '_log2_vs_PGlibrary'] = lrp[c].round(3).values
        s24[c + '_q_vs_PGlibrary'] = qvp[c].round(4).values


    # ---------- sheet 25 : P/G substrates vs P/G non-substrates --------------
    pgA = list(lib[lib.is_PG & lib.is_UBR3_substrate].peptide_24mer)
    pgB = list(lib[lib.is_PG & ~lib.is_UBR3_substrate].peptide_24mer)
    cA, cB = U.class_matrix(pgA), U.class_matrix(pgB)
    lrc2, qc2, _ = U.enrichment_test(pgA, pgB, groups=U.CLASS_MEMBERS)
    lrr2, qr2, _ = U.enrichment_test(pgA, pgB)
    s25 = pd.DataFrame({'position': cA.index})
    for c in U.CLASS_ORDER:
        s25[c + '_substrates_pct'] = (100 * cA[c]).round(2).values
        s25[c + '_nonsubstrates_pct'] = (100 * cB[c]).round(2).values
        s25[c + '_log2'] = lrc2[c].round(3).values
        s25[c + '_q'] = qc2[c].round(4).values
    rows = []
    for pos in qr2.index:
        for a in qr2.columns:
            rows.append({'position': pos, 'residue': a, 'chemical_class': U.AA_CLASS[a],
                         'log2_enrichment': round(lrr2.loc[pos, a], 3),
                         'q_value_BH': float(f'{qr2.loc[pos, a]:.3g}'),
                         'significant_q<0.05': 'yes' if qr2.loc[pos, a] < 0.05 else 'no'})
    s26 = pd.DataFrame(rows).sort_values('q_value_BH')

    sheets = {
        '00_README': readme_sheet(S),
        '01_UBR3_substrates': s01,
        '02_PGED_motif_all': s02,
        '03_PG_library_all': s03,
        '04_substrates_PG_only': s04,
        '05_counts_summary': s05,
        '06_enrich_pos2_residue': s06,
        '07_enrich_motif_class': s07,
        '08_AAclass_by_position': s08,
        '09_posAA_freq_substrates': s09,
        '10_posAA_freq_library': s10,
        '11_posAA_log2_enrichment': s11,
        '12_posAA_fisher_all_cells': s12,
        '13_posAA_qvalues': s13,
        '14_PSI_strata_summary': s14,
        '15_motif_x_PSI_strata': s15,
        '16_substrates_PSI_transition': s16,
        '17_dPSI_by_control_PSI_bin': s17,
        '18_PSI_cutoff_robustness': s18,
        '19_stability_x_substrate': s19,
        '20_PGED_substrates_annotated': s20,
        '21_downstream_residue': s21,
        '22_downstream_class': s22,
        '23_downstream_composition': s23,
        '24_PG_substrate_class_composition': s24,
        '25_PGsub_vs_PGnonsub_class': s25,
        '26_PGsub_vs_PGnonsub_residue': s26,
    }

    with pd.ExcelWriter(OUT, engine='openpyxl') as xw:
        for name, df in sheets.items():
            df.to_excel(xw, name, index=False)
        style(xw.book, sheets)

    print(f'wrote {OUT}')
    for k, v in sheets.items():
        print(f'   {k:28s} {v.shape[0]:>6d} rows x {v.shape[1]:>3d} cols')
    return S


def style(wb, sheets):
    hdr_fill = PatternFill('solid', fgColor='1F4E78')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    # Conditional-format fills live in a <dxf>, and Excel reads bgColor there -
    # not fgColor as it does for ordinary cell styles. Set both so the fill shows.
    def cf_fill(rgb):
        return PatternFill(fill_type='solid', start_color=rgb, end_color=rgb)

    row_fill = cf_fill('FFF8CBAD')                        # tint of the figure orange
    row_font = Font(bold=True, color='FF7A2E0E')
    for name, ws in ((s.title, s) for s in wb.worksheets):
        df = sheets[name]
        ws.freeze_panes = 'B2' if name != '00_README' else 'A2'
        ws.auto_filter.ref = ws.dimensions if name not in ('00_README',) else None
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(1, c)
            cell.fill, cell.font = hdr_fill, hdr_font
            cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='center')
        ws.row_dimensions[1].height = 34
        for c in range(1, ws.max_column + 1):
            col = get_column_letter(c)
            vals = [len(str(ws.cell(r, c).value)) for r in range(2, min(ws.max_row, 300) + 1)
                    if ws.cell(r, c).value is not None]
            head = len(str(ws.cell(1, c).value or ''))
            ws.column_dimensions[col].width = min(max((max(vals) if vals else 8), min(head, 18)) + 2, 46)
        # mark substrates: the flag cell in green, and - on the long library
        # listings - the whole row, so the handful of substrates are findable
        # among thousands of rows without filtering
        cols = list(df.columns)
        if 'UBR3_substrate' in cols:
            L = get_column_letter(cols.index('UBR3_substrate') + 1)
            last = get_column_letter(ws.max_column)
            ws.conditional_formatting.add(
                f'A2:{last}{ws.max_row}',
                FormulaRule(formula=[f'${L}2="yes"'], fill=row_fill, font=row_font, stopIfTrue=False))
        for target in ('mean_dPSI', 'log2_fold', 'fold_enrichment'):
            if target in cols:
                L = get_column_letter(cols.index(target) + 1)
                ws.conditional_formatting.add(
                    f'{L}2:{L}{ws.max_row}',
                    ColorScaleRule(start_type='min', start_color='FFF5F0',
                                   end_type='max', end_color='2A78D6'))
    rm = wb['00_README']
    rm.column_dimensions['A'].width = 30
    rm.column_dimensions['B'].width = 130
    for r in range(1, rm.max_row + 1):
        rm.cell(r, 1).font = Font(bold=True, size=10)
        rm.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')


if __name__ == '__main__':
    main()
